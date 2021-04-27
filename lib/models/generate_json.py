"""
.. module:: generate_json.py
   :license: GPL/CeCIL
   :platform: Unix, Windows
   :synopsis: Generates CORDEXP model documentation JSON documents.

.. moduleauthor:: Mark Conway-Greenslade <momipsl@ipsl.jussieu.fr>

"""
import argparse
import collections
import os

import openpyxl
import pyessv

from lib.utils import io_mgr
from lib.utils import logger
from lib.utils import vocabs



# Define command line argument parser.
_ARGS = argparse.ArgumentParser("Generates CORDEXP model JSON files.")
_ARGS.add_argument(
    "--institution-id",
    help="An institution identifier",
    dest="institution_id",
    type=str
    )

# MIP era.
_MIP_ERA = "cordexp"

# Name of file controlling publication.
_MODEL_PUBLICATION_FNAME = "model_publication.json"


def _main(args):
    """Main entry point.

    """
    for i, s, t, d in vocabs.yield_topics(args.institution_id):
        # print("{} :: {} :: {} :: {}".format(i, s, t, d))
        print(io_mgr.get_model_topic_xls(i, s, t, d))
        # try:
        #     wb = _get_spreadsheet(i, s, t, d)
        # except IOError:
        #     warning = '{} :: {} :: {} :: spreadsheet not found'
        #     warning = warning.format(i.canonical_name, s.canonical_name, t.canonical_name)
        #     logger.log_warning(warning)
        #     continue

        # # Set JSON content.
        # content = _get_content(i, s, t, wb)

        # # Write JSON file.
        # io_mgr.write_model_topic_json(i, s, t, d, content)


def _get_spreadsheet(i, s, t, d):
    """Returns a model topic spreadsheet for processing.

    """
    fpath = io_mgr.get_model_topic_xls(i, s, t, d)
    print(fpath)
    if not os.path.exists(fpath):
        raise IOError()

    return openpyxl.load_workbook(fpath, read_only=True)


def _get_content(i, s, t, wb):
    """Returns content to be written to file system.

    """
    # Initialise output.
    obj = collections.OrderedDict()
    obj['mipEra'] = _MIP_ERA
    obj['institute'] = i.canonical_name
    obj['seedingSource'] = 'Spreadsheet'
    obj['sourceID'] = s.canonical_name
    obj['topic'] = t.canonical_name
    obj['content'] = collections.OrderedDict()

    # Process spreadsheet.
    for idx, ws in enumerate(wb):
        # Process citations/responsible parties.
        if idx == 1:
            # TODO
            continue

        # Extract specialization entries.
        elif idx > 1:
            _set_xls_content(obj, ws)
    
    return obj


def _set_xls_content(obj, ws):
    """Sets content for a particular specialization.

    """
    content = None
    content_is_enum = None
    for row in ws.iter_rows(min_row=1, max_col=4, max_row=ws.max_row):
        if len(row) == 0:
            continue

        # Separation row - begin new specialization block.
        if row[1].value is None:
            if content is not None:
                if content[1]:
                    obj['content'][content[0]] = {'values': content[1]}
                content = None

        # Specialization begin row - spec. id is hidden in 3rd column.
        elif row[2].value is not None:
            content = (row[2].value, [])
            content_is_enum = row[0].value == 'ENUM'

        # Specialization value row.
        elif content is not None and not _is_note(row[1]):
            value = row[1].value

            # Open enums are treated differently.
            if content_is_enum:
                if row[3].value:
                    value = 'Other: ' + row[3].value
                elif value == 'Other: document in cell to the right':
                    break

            # This can occur when spreadsheet is deformatted.
            if value == '=TRUE()':
                value = True
            elif value == '=FALSE()':
                value = False

            # If value is not in blocklist then emit.
            if value not in ('-', 'Other: -'):
                content[1].append(value)

    # Apply last specialization block.
    if content is not None:
        if content[1]:
            obj['content'][content[0]] = {'values': content[1]}


def _is_note(cell):
    """Returns flag indicating whether a cell represents a note tot he user or not.

    """
    try:
        return cell.value.startswith('NOTE: ')
    except AttributeError:
        return False


# Main entry point.
if __name__ == '__main__':
    _main(_ARGS.parse_args())
